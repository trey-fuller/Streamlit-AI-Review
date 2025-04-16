import streamlit as st
import pandas as pd
import numpy as np
import openpyxl
from io import BytesIO
from openpyxl.styles.alignment import Alignment
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="AI Review Workflow", layout="wide")
st.title("AI Review Workflow")

uploaded_file = st.file_uploader("Upload an Excel file", type=["xlsx"])

if uploaded_file:
    xls = pd.ExcelFile(uploaded_file, engine="openpyxl")
    sheet_names = xls.sheet_names

    if "Case Data" not in sheet_names:
        st.error("Sheet 'Case Data' not found in the uploaded file.")
        st.stop()

    if "all_sheets" not in st.session_state:
        all_sheets = {sheet: pd.read_excel(xls, sheet) for sheet in xls.sheet_names}

        for sheet in all_sheets:
            all_sheets[sheet].columns = (
                all_sheets[sheet]
                .columns.map(str)
                .str.strip()
                .str.replace(r'\s+', '_', regex=True)
                .str.lower()
            )

        df = next((v for k, v in all_sheets.items() if k.strip().lower() == "case data"), pd.DataFrame())

        if "completed" in df.columns:
            df["completed"] = df["completed"].fillna("no").astype(str).str.strip().str.lower()
        else:
            df["completed"] = "no"

        index_sheet = all_sheets.get("index", pd.DataFrame())
        if not index_sheet.empty and "sheet" in index_sheet.columns and "last_index" in index_sheet.columns:
            last_index = (
                index_sheet.set_index("sheet").to_dict().get("last_index", {}).get("case_data", 0)
            )
        else:
            last_index = 0

        reviewed_cases = df[df["completed"] == "yes"]
        unreviewed_cases = df[df["completed"] == "no"]

        if not unreviewed_cases.empty:
            current_case_index = unreviewed_cases.index.min()
        else:
            current_case_index = df.index[-1]
            st.balloons()
            st.success("All cases have been completed.")

        st.session_state["current_case_index"] = current_case_index
        st.session_state["all_sheets"] = all_sheets
        st.session_state["df"] = df

    df = st.session_state["df"]
    all_sheets = st.session_state["all_sheets"]
    case_indices = df.index.tolist()
    reviewed_cases = df[df["completed"] == "yes"]
    unreviewed_cases = df[df["completed"] == "no"]
    current_index = st.session_state.get("current_case_index")

    if len(df) == 0:
        st.warning("The uploaded file has no cases to review.")
        st.stop()

    if pd.isna(current_index) or current_index not in df.index:
        if not unreviewed_cases.empty:
            current_index = unreviewed_cases.index.min()
        else:
            current_index = df.index[-1]
            st.balloons()
            st.success("All cases have been completed.")

        st.session_state["current_case_index"] = current_index

    case = df.loc[current_index]

if uploaded_file and "df" in st.session_state and st.session_state["df"] is not None:
    st.write(f"Progress: {len(reviewed_cases)}/{len(df)} cases completed")
    st.write(f"Case {current_index+1}/{len(df)}: {case.get('accession', '')}")

    col_open, col_text = st.columns([1, 2])
    with col_open:
        studio_url = case.get("studio_link", "")
        if studio_url:
            st.markdown(f"<a href='{studio_url}' target='_blank'><button>Open Studio Link</button></a>", unsafe_allow_html=True)

    with col_text:
        st.write("👈 **Click here to launch a case.**")

    default_values = {
        "tp-fp": "TP",
        "second-opinion": False,
        "request-report": "No",
        "location-type": "",
        "comment": ""
    }

    def reset_form(idx):
        row = df.loc[idx]
        if row["completed"] == "no":
            for k, v in default_values.items():
                st.session_state[f"{k}_{idx}"] = v
        else:
            st.session_state[f"tp-fp_{idx}"] = row.get("review_(tp/fp)", "TP")
            st.session_state[f"second-opinion_{idx}"] = row.get("2nd_opinion_(y/n)", "No") == "Yes"
            st.session_state[f"request-report_{idx}"] = row.get("request_report_(y/n)", "No")
            st.session_state[f"location-type_{idx}"] = row.get("location/type", "")
            st.session_state[f"comment_{idx}"] = row.get("comments", "")

    if (
        f"tp-fp_{current_index}" not in st.session_state
        and not st.session_state.get("just_submitted", False)
    ):
        reset_form(current_index)

    if current_index is not None:
        tp_fp = st.radio("True Positive / False Positive", ["TP", "FP"], key=f"tp-fp_{current_index}", index=0 if st.session_state.get(f"tp-fp_{current_index}") == "TP" else 1)
        second_opinion = st.checkbox("Request Second Opinion", key=f"second-opinion_{current_index}", value=st.session_state.get(f"second-opinion_{current_index}", False))
        request_report = st.radio("Request Report", ["No", "Yes"], key=f"request-report_{current_index}", index=0 if st.session_state.get(f"request-report_{current_index}") == "No" else 1)
        location_type = st.text_area("Location/Type", key=f"location-type_{current_index}", value=st.session_state.get(f"location-type_{current_index}", ""))
        comments = st.text_area("Comments (Optional)", key=f"comment_{current_index}", value=st.session_state.get(f"comment_{current_index}", ""))

        col1, col2, col3 = st.columns([1, 2, 1])
        with col1:
            if st.button("Previous Case"):
                if current_index in case_indices:
                    prev_idx = case_indices.index(current_index) - 1
                    if prev_idx >= 0:
                        st.session_state["current_case_index"] = case_indices[prev_idx]
                        reset_form(st.session_state["current_case_index"])
                        st.rerun()

        with col2:
            if st.button("Next Case"):
                if current_index in case_indices:
                    next_idx = case_indices.index(current_index) + 1
                    if next_idx < len(case_indices):
                        st.session_state["current_case_index"] = case_indices[next_idx]
                        reset_form(st.session_state["current_case_index"])
                        st.rerun()

        with col3:
            if st.button("Submit & Next"):
                df.at[current_index, "review_(tp/fp)"] = str(tp_fp)
                df.at[current_index, "2nd_opinion_(y/n)"] = str("Yes" if second_opinion else "No")
                df.at[current_index, "request_report_(y/n)"] = str(request_report)
                df.at[current_index, "location/type"] = str(location_type).strip()
                df.at[current_index, "comments"] = str(comments).strip()
                df.at[current_index, "completed"] = "yes"

                all_sheets["case_data"] = df
                index_sheet = all_sheets.get("index", pd.DataFrame(columns=["sheet", "last_index"]))
                index_sheet.columns = index_sheet.columns.str.strip().str.lower()
                if "sheet" not in index_sheet.columns or "last_index" not in index_sheet.columns:
                    index_sheet = pd.DataFrame(columns=["sheet", "last_index"])
                if not index_sheet.empty:
                    index_sheet = index_sheet.set_index("sheet")
                index_sheet.loc["case_data", "last_index"] = current_index
                all_sheets["index"] = index_sheet.reset_index()

                st.session_state["df"] = df
                st.session_state["all_sheets"] = all_sheets

                reviewed_cases = df[df["completed"] == "yes"]
                unreviewed_cases = df[df["completed"] == "no"]
                next_unreviewed = unreviewed_cases.index[unreviewed_cases.index > current_index].min() if not unreviewed_cases.empty else None
                st.session_state["current_case_index"] = next_unreviewed if not pd.isna(next_unreviewed) else None
                st.session_state["just_submitted"] = True

                if st.session_state["current_case_index"] is not None:
                    next_case = df.loc[st.session_state["current_case_index"]]
                    next_studio_url = next_case.get("studio_link", "")
                    if next_studio_url:
                        st.markdown(f"<script>window.open('{next_studio_url}', '_blank');</script>", unsafe_allow_html=True)

                    reset_form(st.session_state["current_case_index"])
                else:
                    st.balloons()
                    st.success("You have completed all available cases!")

                st.rerun()

    tab1, tab2 = st.tabs(["Case Review", "Login Info"])
    with tab2:
        st.subheader("Login Info")
        st.write("**Username:** rpxuser")
        st.write("**Password:** PpD4u2RK")

    st.markdown("---")
    st.subheader("Download Completed Workbook")

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet, data in all_sheets.items():
            data.to_excel(writer, sheet_name=sheet, index=False)
            ws = writer.sheets[sheet]
            for i, column in enumerate(data.columns, 1):
                ws.column_dimensions[get_column_letter(i)].width = 20
                for cell in ws[get_column_letter(i)]:
                    cell.alignment = Alignment(horizontal="left")

    st.download_button(
        label="📥 Download Updated Excel",
        data=output.getvalue(),
        file_name=f"{uploaded_file.name.replace('.xlsx', '')}-updated-{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
